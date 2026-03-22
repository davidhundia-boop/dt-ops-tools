"""
Campaign optimizer backend — Digital Turbine preload pipeline.
Implements full rules: data prep, segmentation, progression, discard rules,
daily cap logic, bid optimization, and color-coded Excel output.

NOTE: "Optimization Suggestions" refers to activating this optimizer.py module only.
- site_performance = internal file (Excel .xlsx)
- DT_DX = client file with performance data (CSV)
"""

from io import BytesIO
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# --- Column name normalization (internal file may use camelCase or snake_case) ---
def _norm_col(s):
    if not isinstance(s, str):
        return s
    return str(s).strip()


def _find_col(df, *candidates):
    """Return first column name in df that matches any candidate (case-insensitive)."""
    cols = {_norm_col(c).lower(): c for c in df.columns}
    for cand in candidates:
        k = _norm_col(cand).lower()
        if k in cols:
            return cols[k]
        # also try exact match after strip
        for c in df.columns:
            if _norm_col(c).lower() == k:
                return c
    return None


def col_letter_to_idx(letter):
    """Convert Excel column letter (A–Z, AA–AZ, etc.) to 0-based index. A=0, B=1, …, Z=25, AA=26, AB=27, …"""
    letter = str(letter).strip().upper()
    if not letter or not all(c.isalpha() for c in letter):
        raise ValueError(f"Column must be letters A–Z or AA–AZ etc., got: {letter!r}")
    idx = 0
    for c in letter:
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1


def find_col_by_pattern(df, pattern):
    """
    Find a column in df that contains the given pattern (case-insensitive).
    Returns the column name if found, else None.
    """
    pattern_lower = pattern.lower().strip()
    for col in df.columns:
        if pattern_lower in str(col).lower():
            return col
    return None


def col_name_or_letter_to_idx(df, col_spec):
    """
    Convert column specification to 0-based index.
    col_spec can be:
      - A single letter (A–Z) or multi-letter (AA–AZ) for Excel-style column index
      - A column name or partial name to search for
    Returns (index, column_name) tuple.
    """
    col_spec = str(col_spec).strip()
    
    # Check if it looks like an Excel column letter (all letters, 1-3 chars)
    if col_spec.isalpha() and len(col_spec) <= 3:
        idx = col_letter_to_idx(col_spec)
        if idx < len(df.columns):
            return idx, df.columns[idx]
        raise ValueError(f"Column letter {col_spec} is out of range (file has {len(df.columns)} columns)")
    
    # Otherwise, search by name pattern
    col_name = find_col_by_pattern(df, col_spec)
    if col_name:
        return df.columns.get_loc(col_name), col_name
    
    raise ValueError(f"Could not find column matching '{col_spec}' in the file")


def _parse_pct(val):
    """Parse percentage string (e.g. '5.9%') or number to decimal (0.059)."""
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float)):
        if val > 1:
            return val / 100.0
        return float(val)
    s = str(val).strip()
    if not s:
        return np.nan
    s = s.replace("%", "").replace(",", ".").strip()
    try:
        x = float(s)
        return x / 100.0 if x > 1 else x
    except ValueError:
        return np.nan


def _parse_roas(val):
    """
    Parse ROAS value (e.g. '2.18%', '2.18', '218%') to a ratio.
    ROAS is typically expressed as a ratio (e.g., 2.18 means $2.18 return per $1 spent).
    
    Rules:
    - If value has '%' and is > 1 after removing %, treat as percentage (218% -> 2.18)
    - If value has '%' and is <= 1 after removing %, treat as actual percentage (2.18% -> 0.0218)
    - If numeric and > 10, assume it's a percentage representation (218 -> 2.18)
    - If numeric and <= 10, assume it's already a ratio (2.18 -> 2.18)
    """
    if pd.isna(val):
        return np.nan
    
    s = str(val).strip()
    if not s:
        return np.nan
    
    has_pct = '%' in s
    s = s.replace("%", "").replace(",", ".").strip()
    
    try:
        x = float(s)
    except ValueError:
        return np.nan
    
    if has_pct:
        # If it had a % sign, convert from percentage to ratio
        return x / 100.0
    else:
        # No % sign - if > 10, assume it's percentage representation
        if x > 10:
            return x / 100.0
        return x


# --- Excluded site types ---
EXCLUDED_PATTERNS = [
    "om push", "om_push", "notifications",
]


def _is_excluded(name):
    if pd.isna(name):
        return False
    n = str(name).lower()
    return any(p in n for p in EXCLUDED_PATTERNS)


# --- Segment colors ---
SEGMENT_COLORS = {
    "green": "C6EFCE",
    "yellow": "FFEB9C",
    "orange": "FFCC99",
    "red": "FFC7CE",
}
DAILY_CAP_FILL = "DAE3F3"
HEADER_FILL = "1F3864"


def _load_internal(path):
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [_norm_col(c) for c in df.columns]
    return df


def _load_advertiser(path, kpi_d7_idx, kpi_d2nd_idx):
    df = pd.read_csv(path, encoding="utf-8", encoding_errors="replace")
    df.columns = [_norm_col(c) for c in df.columns]
    return df


def _ensure_key(df, campaign_col, site_id_col):
    """Create Key = campaignName + '_' + siteId (integer string)."""
    def to_str(x):
        if pd.isna(x):
            return ""
        if isinstance(x, float) and x == int(x):
            return str(int(x))
        return str(int(float(x))) if isinstance(x, (int, float)) else str(x)
    df["Key"] = df[campaign_col].astype(str) + "_" + df[site_id_col].apply(to_str)
    return df


def run_optimization(
    internal_file,
    advertiser_file,
    kpi_col_d7_idx=None,
    kpi_col_d2nd_idx=None,
    kpi_d7_pct=None,
    kpi_d2nd_pct=None,
    weight_main=0.80,
    weight_secondary=0.20,
    kpi_col_d7_spec=None,
    kpi_col_d2nd_spec=None,
    kpi_mode="roi",
):
    """
    Run the full optimization pipeline.
    Returns (output_bytes: BytesIO, summary: dict).
    
    Parameters:
    - internal_file: Path to internal campaign data (Excel .xlsx)
    - advertiser_file: Path to advertiser performance report (CSV)
    - kpi_col_d7_idx: (deprecated) 0-based column index for D7 KPI
    - kpi_col_d2nd_idx: (deprecated) 0-based column index for secondary KPI
    - kpi_d7_pct: Target for D7 KPI (as percentage, e.g., 2.18 for 2.18%)
    - kpi_d2nd_pct: Target for secondary KPI (as percentage)
    - weight_main: Weight for D7 KPI (0-1)
    - weight_secondary: Weight for secondary KPI (0-1)
    - kpi_col_d7_spec: Column letter (e.g., 'I') or name pattern (e.g., 'ROAS D7') for D7 KPI
    - kpi_col_d2nd_spec: Column letter or name pattern for secondary KPI
    - kpi_mode: 'roi' (percentage-based) or 'roas' (ratio-based)
    """
    weight_main = float(weight_main)
    weight_secondary = float(weight_secondary)
    if abs(weight_main + weight_secondary - 1.0) > 0.01:
        weight_main, weight_secondary = 0.80, 0.20

    internal = _load_internal(internal_file)
    advertiser = _load_advertiser(advertiser_file, kpi_col_d7_idx, kpi_col_d2nd_idx)
    
    # Resolve column specifications to indices
    if kpi_col_d7_spec is not None:
        kpi_col_d7_idx, d7_col_resolved = col_name_or_letter_to_idx(advertiser, kpi_col_d7_spec)
    if kpi_col_d2nd_spec is not None:
        kpi_col_d2nd_idx, d2nd_col_resolved = col_name_or_letter_to_idx(advertiser, kpi_col_d2nd_spec)

    # --- Internal: column names ---
    campaign_col = _find_col(internal, "campaignName", "campaign_name")
    site_id_col = _find_col(internal, "siteId", "site_id")
    site_name_col = _find_col(internal, "siteName", "site_name")
    if not campaign_col or not site_id_col:
        raise ValueError("Internal file must contain campaignName and siteId (or campaign_name, site_id).")
    if not site_name_col:
        site_name_col = "siteName"

    # Step 1: Exclude site types
    internal = internal[
        ~internal[campaign_col].apply(_is_excluded) & ~internal[site_name_col].apply(_is_excluded)
    ].copy()

    _ensure_key(internal, campaign_col, site_id_col)

    # Advertiser: build Key and KPI columns
    adv_campaign = _find_col(advertiser, "campaignName", "campaign_name")
    adv_site_id = _find_col(advertiser, "siteId", "site_id")
    if not adv_campaign or not adv_site_id:
        raise ValueError("Advertiser file must contain campaignName and siteId (or equivalents).")
    _ensure_key(advertiser, adv_campaign, adv_site_id)

    # KPI columns by index (CSV columns)
    ncols = len(advertiser.columns)
    if kpi_col_d7_idx >= ncols or kpi_col_d2nd_idx >= ncols:
        raise ValueError("KPI column indices out of range for advertiser file.")
    d7_col_name = advertiser.columns[kpi_col_d7_idx]
    d2nd_col_name = advertiser.columns[kpi_col_d2nd_idx]
    roi_d2nd_label = d2nd_col_name  # e.g. "ROI D30" or "ROI D14"

    # Parse KPI values based on mode
    if kpi_mode == "roas":
        advertiser["ROI D7"] = advertiser.iloc[:, kpi_col_d7_idx].apply(_parse_roas)
        advertiser["ROI D2nd"] = advertiser.iloc[:, kpi_col_d2nd_idx].apply(_parse_roas)
    else:
        advertiser["ROI D7"] = advertiser.iloc[:, kpi_col_d7_idx].apply(_parse_pct)
        advertiser["ROI D2nd"] = advertiser.iloc[:, kpi_col_d2nd_idx].apply(_parse_pct)
    merge_cols = ["Key", "ROI D7", "ROI D2nd"]
    advertiser_merge = advertiser[merge_cols].drop_duplicates(subset=["Key"], keep="first")

    internal = internal.merge(advertiser_merge, on="Key", how="left", suffixes=("", "_y"))

    # Drop N/A: ROI D7, ROI D2nd, maxPreloads, fillRate
    max_preloads_col = _find_col(internal, "maxPreloads", "max_preloads")
    fill_rate_col = _find_col(internal, "fillRate", "fill_rate")
    if max_preloads_col and fill_rate_col:
        internal = internal.dropna(subset=["ROI D7", "ROI D2nd", max_preloads_col, fill_rate_col])
    else:
        internal = internal.dropna(subset=["ROI D7", "ROI D2nd"])
        if max_preloads_col:
            internal = internal.dropna(subset=[max_preloads_col])
        if fill_rate_col:
            internal = internal.dropna(subset=[fill_rate_col])

    # Normalize fill rate to 0-1 if it's 0-100
    if fill_rate_col:
        fr = internal[fill_rate_col]
        if fr.max() > 1.5:
            internal[fill_rate_col] = fr / 100.0
    fill_rate_col = fill_rate_col or "fillRate"
    if fill_rate_col not in internal.columns:
        internal[fill_rate_col] = np.nan

    # Standardize names for rest of pipeline
    renames = {}
    for c in internal.columns:
        if _norm_col(c).lower() == "campaignid": renames[c] = "campaignId"
        if _norm_col(c).lower() == "campaignname": renames[c] = "campaignName"
        if _norm_col(c).lower() == "siteid": renames[c] = "siteId"
        if _norm_col(c).lower() == "sitename": renames[c] = "siteName"
        if _norm_col(c).lower() == "maxpreloads": renames[c] = "maxPreloads"
        if _norm_col(c).lower() == "fillrate": renames[c] = "fillRate"
        if _norm_col(c).lower() == "effectivebidfloor": renames[c] = "effectiveBidFloor"
        if _norm_col(c).lower() == "bidrate": renames[c] = "bidRate"
        if _norm_col(c).lower() == "dailycap": renames[c] = "dailyCap"
        if _norm_col(c).lower() == "lowtier": renames[c] = "lowTier"
        if _norm_col(c).lower() == "midtier": renames[c] = "midTier"
        if _norm_col(c).lower() == "hightier": renames[c] = "highTier"
        if _norm_col(c).lower() == "bidfloorgroupname": renames[c] = "bidFloorGroupName"
        if _norm_col(c).lower() == "ecpp": renames[c] = "ecpp"
        if _norm_col(c).lower() == "ecpi": renames[c] = "ecpi"
        if _norm_col(c).lower() == "cvr": renames[c] = "cvr"
    internal = internal.rename(columns=renames)
    campaign_col = "campaignName" if "campaignName" in internal.columns else campaign_col
    site_id_col = "siteId" if "siteId" in internal.columns else site_id_col
    site_name_col = "siteName" if "siteName" in internal.columns else site_name_col

    # Ensure numeric columns and calculate target based on mode
    if kpi_mode == "roas":
        # For ROAS mode, the target is entered as percentage (e.g., 2.18 means 2.18% = 0.0218)
        kpi_d7 = float(kpi_d7_pct) / 100.0
        kpi_d2nd = float(kpi_d2nd_pct) / 100.0 if kpi_d2nd_pct else kpi_d7
    else:
        # For ROI mode, target is percentage (e.g., 10 means 10% = 0.10)
        kpi_d7 = float(kpi_d7_pct) / 100.0
        kpi_d2nd = float(kpi_d2nd_pct) / 100.0 if kpi_d2nd_pct else kpi_d7
    target = weight_main * kpi_d7 + weight_secondary * kpi_d2nd

    internal["score"] = weight_main * internal["ROI D7"].astype(float) + weight_secondary * internal["ROI D2nd"].astype(float)
    internal["target"] = target
    internal["pct_below"] = np.where(
        internal["target"] > 0,
        (internal["target"] - internal["score"]) / internal["target"],
        np.nan,
    )
    internal["pct_above"] = np.where(
        internal["target"] > 0,
        (internal["score"] - internal["target"]) / internal["target"],
        np.nan,
    )

    # Segment
    def segment(row):
        roi_d7 = row["ROI D7"]
        roi_d2nd = row["ROI D2nd"]
        sc = row["score"]
        tg = row["target"]
        pct_below = row["pct_below"]
        if roi_d7 == 0 and roi_d2nd == 0:
            return "red"
        if sc >= tg:
            return "green"
        if pd.isna(pct_below) or pct_below <= 0:
            return "yellow"
        if pct_below <= 0.50:
            return "yellow"
        if pct_below < 1.0:
            return "orange"
        return "orange"

    internal["segment"] = internal.apply(segment, axis=1)

    # Progression
    def progression(row):
        d7 = row["ROI D7"]
        d2nd = row["ROI D2nd"]
        if d7 == 0:
            return "flat"
        if d2nd > d7:
            return "good"
        if d2nd < d7:
            return "poor"
        return "flat"

    internal["progression"] = internal.apply(progression, axis=1)

    # Spend, preloads, installs, status
    spend_col = _find_col(internal, "spend", "Spend")
    preloads_col = _find_col(internal, "preloads", "Preloads")
    installs_col = _find_col(internal, "installs", "Installs")
    status_col = _find_col(internal, "status", "Status")
    if not spend_col:
        internal["_spend"] = 0.0
        spend_col = "_spend"
    else:
        internal[spend_col] = pd.to_numeric(internal[spend_col], errors="coerce").fillna(0)
    if not preloads_col:
        internal["_preloads"] = 0
        preloads_col = "_preloads"
    else:
        internal[preloads_col] = pd.to_numeric(internal[preloads_col], errors="coerce").fillna(0)
    if not installs_col:
        internal["_installs"] = 0
        installs_col = "_installs"
    else:
        internal[installs_col] = pd.to_numeric(internal[installs_col], errors="coerce").fillna(0)
    if not status_col:
        internal["_status"] = ""
        status_col = "_status"
    else:
        internal[status_col] = internal[status_col].astype(str).str.strip().str.lower()

    # Discard flag
    def should_discard(row):
        spend = float(row[spend_col]) if spend_col in row else 0
        preloads = float(row[preloads_col]) if preloads_col in row else 0
        inst = int(row[installs_col]) if installs_col in row else 0
        seg = row["segment"]
        status = (row[status_col] or "") if status_col in row else ""
        prog = row["progression"]
        fill_rate = row[fill_rate_col] if fill_rate_col in row else 0
        roi_d7 = row["ROI D7"]
        roi_d2nd = row["ROI D2nd"]

        if seg == "green" and inst >= 5:
            return False
        if prog == "good" and fill_rate < 0.60 and roi_d7 > 0 and spend >= 100:
            return False
        if spend < 100:
            return True
        if preloads < 100:
            return True
        if status == "paused" and seg != "green":
            return True
        return False

    internal["_discard"] = internal.apply(should_discard, axis=1)

    # Daily cap suggestion (Step 5)
    bid_floor_col = _find_col(internal, "effectiveBidFloor", "effective_bid_floor") or "effectiveBidFloor"
    bid_rate_col = _find_col(internal, "bidRate", "bid_rate") or "bidRate"
    if bid_floor_col not in internal.columns:
        internal[bid_floor_col] = np.nan
    if bid_rate_col not in internal.columns:
        internal[bid_rate_col] = np.nan
    internal["bid_floor_num"] = pd.to_numeric(internal[bid_floor_col], errors="coerce")
    internal["bid_rate_num"] = pd.to_numeric(internal[bid_rate_col], errors="coerce")
    high_tier_col = _find_col(internal, "highTier", "high_tier") or "highTier"
    if high_tier_col not in internal.columns:
        internal[high_tier_col] = np.nan
    internal["high_tier_num"] = pd.to_numeric(internal[high_tier_col], errors="coerce")

    def daily_cap_suggestion(row):
        spend = float(row[spend_col])
        if spend <= 1000:
            return ""
        roi_d7 = row["ROI D7"]
        roi_d2nd = row["ROI D2nd"]
        bid_floor = row["bid_floor_num"]
        bid_rate = row["bid_rate_num"]
        at_or_below_floor = (
            not pd.isna(bid_floor) and not pd.isna(bid_rate) and float(bid_rate) <= float(bid_floor)
        )
        if not at_or_below_floor:
            return ""
        has_perf = (roi_d7 and float(roi_d7) > 0) or (roi_d2nd and float(roi_d2nd) > 0)
        if has_perf:
            cap = round((spend / 30) * 0.50, 2)
            return f"Add daily cap ${cap:.2f}"
        return "Suggest pause"

    internal["Daily Cap Suggestion"] = internal.apply(daily_cap_suggestion, axis=1)

    # Bid logic
    def action_and_bid(row):
        if row["_discard"]:
            return "", np.nan
        spend = float(row[spend_col])
        fill_rate = float(row[fill_rate_col]) if fill_rate_col in row and pd.notna(row[fill_rate_col]) else 0
        roi_d7 = row["ROI D7"]
        roi_d2nd = row["ROI D2nd"]
        seg = row["segment"]
        prog = row["progression"]
        pct_below = row["pct_below"]
        pct_above = row["pct_above"]
        bid_rate = row["bid_rate_num"]
        bid_floor = row["bid_floor_num"]
        high_tier = row["high_tier_num"]
        inst = int(row[installs_col]) if installs_col in row else 0

        if pd.isna(bid_rate):
            bid_rate = 0.0
        else:
            bid_rate = float(bid_rate)
        if pd.isna(bid_floor):
            bid_floor = 0.0
        else:
            bid_floor = float(bid_floor)
        if pd.isna(high_tier):
            high_tier = None
        else:
            high_tier = float(high_tier)

        at_or_below_floor = bid_rate <= bid_floor
        if at_or_below_floor:
            return "", np.nan

        def meet_floor(new_bid):
            if new_bid < bid_floor:
                return "Meet bid floor", round(bid_floor, 2)
            return None, round(new_bid, 2)

        # 6A Good progression
        if prog == "good" and fill_rate < 0.60 and roi_d7 and float(roi_d7) > 0:
            ratio = float(roi_d2nd) / float(roi_d7) if roi_d7 else 0
            inc_pct = 0.15 if ratio >= 2.0 else 0.10
            new_bid = bid_rate * (1 + inc_pct)
            if high_tier is not None and new_bid > high_tier:
                new_bid = high_tier
            act, rec = meet_floor(new_bid)
            if act:
                return act, rec
            return f"Increase bid {int(inc_pct*100)}%", rec

        # 6B Poor progression
        if prog == "poor":
            if spend < 100:
                return "", np.nan
            kpi_d2nd_val = kpi_d2nd
            d2nd_green = roi_d2nd >= kpi_d2nd_val if roi_d2nd is not None else False
            if d2nd_green:
                return "", np.nan
            new_bid = bid_rate * 0.90
            act, rec = meet_floor(new_bid)
            if act:
                return act, rec
            return "Decrease bid 10%", rec

        # 6C Green
        if seg == "green":
            if inst < 5:
                return "", np.nan
            if fill_rate > 0.80 or (0.60 <= fill_rate <= 0.80):
                inc_pct = 0.15
            else:
                if pd.isna(pct_above) or pct_above <= 0.25:
                    inc_pct = 0.10
                elif pct_above <= 0.50:
                    inc_pct = 0.20
                else:
                    inc_pct = 0.30
            new_bid = bid_rate * (1 + inc_pct)
            if high_tier is not None and new_bid > high_tier:
                new_bid = high_tier
            if bid_rate > high_tier and fill_rate < 0.70 and high_tier is not None:
                inc_pct = min(inc_pct, 0.15)
                new_bid = min(bid_rate * (1 + inc_pct), high_tier) if high_tier else bid_rate * (1 + inc_pct)
            act, rec = meet_floor(new_bid)
            if act:
                return act, rec
            return f"Increase bid {int(inc_pct*100)}%", rec

        # 6D Yellow
        if seg == "yellow":
            if pd.isna(pct_below) or pct_below <= 0.25:
                dec = 0.10
            else:
                dec = 0.15
            new_bid = bid_rate * (1 - dec)
            act, rec = meet_floor(new_bid)
            if act:
                return act, rec
            return f"Decrease bid {int(dec*100)}%", rec

        # 6E Orange
        if seg == "orange":
            if pd.isna(pct_below) or pct_below <= 0.75:
                dec = 0.20
            else:
                dec = 0.25
            new_bid = bid_rate * (1 - dec)
            act, rec = meet_floor(new_bid)
            if act:
                return act, rec
            return f"Decrease bid {int(dec*100)}%", rec

        # 6F Red
        if seg == "red":
            new_bid = bid_rate * 0.70
            act, rec = meet_floor(new_bid)
            if act:
                return act, rec
            return "Decrease bid 30%", rec

        return "", np.nan

    internal["Action"] = ""
    internal["Recommended bid"] = np.nan
    for i in internal.index:
        act, rec = action_and_bid(internal.loc[i])
        internal.at[i, "Action"] = act
        internal.at[i, "Recommended bid"] = rec

    # Determine KPI column labels based on mode
    kpi_d7_label = "ROAS D7" if kpi_mode == "roas" else "ROI D7"
    kpi_d2nd_label = "ROAS D2nd" if kpi_mode == "roas" else "ROI D2nd"
    
    # Output columns order
    out_cols = [
        "Key", "campaignId", "campaignName", "siteId", "siteName", "status", "spend", "preloads",
        "maxPreloads", "fillRate", "installs", "cvr", "ecpp", "ecpi", "bidFloorGroupName",
        "effectiveBidFloor", "bidRate", "dailyCap", "lowTier", "midTier", "highTier",
        kpi_d7_label, kpi_d2nd_label, "Action", "Recommended bid", "Daily Cap Suggestion",
    ]
    # Map from our names to output; use first available
    id_col = _find_col(internal, "campaignId", "campaign_id") or campaign_col
    sn_col = _find_col(internal, "siteName", "site_name") or site_name_col
    sid_col = _find_col(internal, "siteId", "site_id") or site_id_col
    maxp_col = _find_col(internal, "maxPreloads", "max_preloads")
    cvr_col = _find_col(internal, "cvr", "CVR")
    ecpp_col = _find_col(internal, "ecpp", "ECPP")
    ecpi_col = _find_col(internal, "ecpi", "ECPI")
    bfg_col = _find_col(internal, "bidFloorGroupName", "bid_floor_group_name")
    dc_col = _find_col(internal, "dailyCap", "daily_cap")
    lt_col = _find_col(internal, "lowTier", "low_tier")
    mt_col = _find_col(internal, "midTier", "mid_tier")
    ht_col = _find_col(internal, "highTier", "high_tier")

    def _get(df, row, col, default=""):
        if col and col in df.columns and col in row:
            v = row[col]
            return v if pd.notna(v) else default
        return default

    result_rows = []
    for _, row in internal.iterrows():
        result_rows.append({
            "Key": row["Key"],
            "campaignId": _get(internal, row, id_col),
            "campaignName": row[campaign_col],
            "siteId": _get(internal, row, sid_col),
            "siteName": row[sn_col],
            "status": _get(internal, row, status_col),
            "spend": row[spend_col],
            "preloads": row[preloads_col],
            "maxPreloads": _get(internal, row, maxp_col),
            "fillRate": row[fill_rate_col],
            "installs": row[installs_col],
            "cvr": _get(internal, row, cvr_col),
            "ecpp": _get(internal, row, ecpp_col),
            "ecpi": _get(internal, row, ecpi_col),
            "bidFloorGroupName": _get(internal, row, bfg_col),
            "effectiveBidFloor": row[bid_floor_col],
            "bidRate": row[bid_rate_col],
            "dailyCap": _get(internal, row, dc_col),
            "lowTier": _get(internal, row, lt_col),
            "midTier": _get(internal, row, mt_col),
            "highTier": _get(internal, row, ht_col),
            kpi_d7_label: row["ROI D7"],
            kpi_d2nd_label: row["ROI D2nd"],
            "Action": row["Action"],
            "Recommended bid": row["Recommended bid"],
            "Daily Cap Suggestion": row["Daily Cap Suggestion"],
            "_segment": row["segment"],
            "_discard": row["_discard"],
        })
    out_df = pd.DataFrame(result_rows)

    # Summary counts
    total_rows = len(out_df)
    rows_actioned = (out_df["Action"] != "").sum()
    rows_disregarded = out_df["_discard"].sum()
    rows_with_cap = (out_df["Daily Cap Suggestion"] != "").sum()
    action_breakdown = out_df[out_df["Action"] != ""]["Action"].value_counts().to_dict()
    segment_breakdown = out_df["_segment"].value_counts().to_dict()

    # Build Excel with formatting
    buf = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Optimization"

    # Drop helper columns for export
    export_df = out_df.drop(columns=["_segment", "_discard"], errors="ignore")
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            else:
                cell.font = Font(name="Arial", size=9)
                seg = out_df.iloc[r_idx - 2]["_segment"] if r_idx > 1 and "_segment" in out_df.columns else None
                discarded = out_df.iloc[r_idx - 2]["_discard"] if r_idx > 1 and "_discard" in out_df.columns else True
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                col_name = export_df.columns[c_idx - 1] if c_idx <= len(export_df.columns) else ""
                if col_name == "Daily Cap Suggestion":
                    cell.fill = PatternFill(start_color=DAILY_CAP_FILL, end_color=DAILY_CAP_FILL, fill_type="solid")
                    cell.font = Font(name="Arial", size=9, bold=True)
                elif not discarded and seg and col_name in (kpi_d7_label, kpi_d2nd_label, "Action", "Recommended bid"):
                    color = SEGMENT_COLORS.get(seg, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                elif not discarded and seg:
                    color = SEGMENT_COLORS.get(seg, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Column widths
    for c_idx, col in enumerate(export_df.columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 14

    # Number formats
    money_cols = ["spend", "ecpp", "ecpi", "effectiveBidFloor", "bidRate", "Recommended bid", "lowTier", "midTier", "highTier"]
    pct_cols = ["fillRate", "cvr", kpi_d7_label, kpi_d2nd_label]
    int_cols = ["preloads", "maxPreloads", "installs", "dailyCap"]
    for r in range(2, ws.max_row + 1):
        for c_idx, col_name in enumerate(export_df.columns, 1):
            cell = ws.cell(row=r, column=c_idx)
            if col_name in money_cols and cell.value is not None and cell.value != "":
                try:
                    v = float(cell.value)
                    cell.number_format = '"$"#,##0.00'
                    cell.value = v
                except (TypeError, ValueError):
                    pass
            elif col_name in pct_cols and cell.value is not None and cell.value != "":
                try:
                    v = float(cell.value)
                    if v <= 1:
                        cell.number_format = '0.00%'
                    else:
                        cell.value = v / 100.0
                        cell.number_format = '0.00%'
                except (TypeError, ValueError):
                    pass
            elif col_name in int_cols and cell.value is not None and cell.value != "":
                try:
                    cell.value = int(float(cell.value))
                    cell.number_format = "#,##0"
                except (TypeError, ValueError):
                    pass
            if col_name in ("campaignId", "siteId"):
                cell.number_format = "@"

    wb.save(buf)
    buf.seek(0)

    summary = {
        "total_rows": int(total_rows),
        "rows_actioned": int(rows_actioned),
        "rows_disregarded": int(rows_disregarded),
        "rows_with_cap": int(rows_with_cap),
        "kpi_d7_col": d7_col_name,
        "kpi_d2nd_col": roi_d2nd_label,
        "kpi_mode": kpi_mode,
        "kpi_d7_target": kpi_d7,
        "kpi_d2nd_target": kpi_d2nd,
        "action_breakdown": action_breakdown,
        "segment_breakdown": segment_breakdown,
    }
    return buf, summary
